import calendar
from datetime import date, timedelta,datetime
import os
import shutil
from urllib import request
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.shortcuts import render,redirect
from django.http import FileResponse, HttpResponse
from openpyxl import load_workbook 
from openpyxl.styles import Border,Side
from pointage.forms import *
from django.forms import formset_factory,inlineformset_factory
from django.contrib.auth import logout
from .models import *
from .utils import *
from shutil import copyfile
from django.contrib.admin.models import ADDITION, CHANGE, DELETION

import time

def pointage2(request,ID):

    unite = Unite.objects.get(id=ID)
    instances=Employe.objects.filter(unite=unite)
    mission_station=Mission.objects.filter(current_unite_id=ID,start_date__lte=datetime.now().date())
    employe_on_mission=Employe.objects.filter(mission__in=mission_station).distinct()
    employe_on_mission=employe_on_mission.exclude(unite_id=ID)
    mission_for_employe_to_disable=Mission.objects.filter(employe__in=instances,start_date__lte=datetime.now().date()).distinct()
    employe_to_disable=Employe.objects.filter(mission__in=mission_for_employe_to_disable,unite=unite).distinct()
    all_employe=list(instances)+list(employe_on_mission)
    date=datetime.now().date()  
    f_date = date - timedelta(days=5)
    date_range = [f_date + timedelta(days=i) for i in range(10)]
    if request.method == 'POST':
        
        MI=['1',"2","3","4","5","MS"]
        Conge=['C','CSS']
        for i in all_employe:
            for date in date_range:
                choice=request.POST.get(f'{i.id}_{date}')
                code=Code_Employe.objects.filter(employe=i,date=date)
                if  choice == "" :
                    if code.exists():
                                month_emp=Month_stat.objects.filter(employe=i,period=f'{date.month}{date.year}')
                                if code.code_id in MI:
                                    month_emp.mission-=1
                                elif code.code_id in "8":
                                    month_emp.absent-=1
                                elif code.code_id in Conge:
                                    month_emp.conge-=1
                                elif code.code_id  == "7":
                                    month_emp.abs_autorise-=1
                                elif code.code_id in "MLD":
                                    month_emp.mld-=1
                                elif code.code_id in "RS":
                                    month_emp.rs-=1
                                elif code.code_id in "6":
                                    month_emp.eve_fam-=1

                                code.delete()
                    else:
                        continue
                elif choice == None :
                    continue
                else:
                    if code.exists():
                        period_i=f'{date.month}{date.year}'
                        try:
                            if code[0].stored:
                                month_emp=Month_stat.objects.get(employe=i,period=period_i)
                                if code[0].code_id in MI:
                                    month_emp.mission-=1
                                elif code[0].code_id in "8":
                                    month_emp.absent-=1
                                elif code[0].code_id in Conge:
                                    month_emp.conge-=1
                                elif code[0].code_id  == "7":
                                    month_emp.abs_autorise-=1
                                elif code[0].code_id in "MLD":
                                    month_emp.mld-=1
                                elif code[0].code_id in "RS":
                                    month_emp.rs-=1
                                elif code[0].code_id in "6":
                                    month_emp.eve_fam-=1
                                elif code[0].code_id in 'T':
                                    month_emp.travail-=1
                        except:
                            pass            
                        code[0].delete()
                        
                    code_emp=Code_Employe()
                    code_emp.employe=i
                    code_emp.date=date
                    code_emp.last_update = timezone.now().date()
                    code_emp.code=Code.objects.get(pk=choice)
                    
                    code_emp.save()
                    
                    log_action(request.user,ContentType.objects.get_for_model(Code_Employe),code_emp.id,str(code_emp),ADDITION,"to be saved")
                    create_log(request.user.username,"add",f"code employe {code_emp}")

            for emp in employe_on_mission:
                valid=request.POST(f"{emp.id}_valid")
                if valid:
                    mission=Mission.objects.get(employe=emp)
                    unites=mission.unites
                    mission.current_unite=unites.split()[1]
                    list_of_unites=unites.split() #this two line to remove the first element 
                    new_unites=' '.join(list_of_unites[1:])
                    mission.unites=new_unites
                    mission.total_validation-=1
                    if mission.total_validation==0:
                        mission.status=2
                        create_log(request.user.username,"valid",f"{mission}")
                    
            unite.last_update= datetime.now().date()
            unite.save()

        return redirect("menu_view")
    else:
        if date.day>=16:
            start_date=datetime(date.year,date.month,16)
            end_date=datetime(date.year,date.month+1,15)
            if end_date.month==13:
                end_date.month=1
        else:
            start_date=datetime(date.year,date.month-1,16)
            end_date=datetime(date.year,date.month,15)
            if start_date.month==0:
                start_date.month=12
        tmp = {}
        res = {}
        overworked_employes=[]
        for i in instances :
                codes_emp = Code_Employe.objects.filter(employe=i)
                overwork=codes_emp.filter(code_id="T",date__range=(start_date,end_date)).count()
                if overwork > 28 :
                    overworked_employes.append(i)  
                tmp = {}
                for d in date_range:
                    try :
                        e = codes_emp.get(date=d)
                        
                        tmp[str(d)] = e 
                    except:
                        tmp[str(d)] = ""
                
                res[i.id] = tmp
        
        
        res2 = {}
        for i in employe_on_mission :
                codes_emp = Code_Employe.objects.filter(employe=i)
                overwork=codes_emp.filter(code_id="T",date__range=(start_date,end_date)).count()
                if overwork > 28 :
                    overworked_employes.append(i)  
                tmp = {}
                for d in date_range:
                    try :
                        e = codes_emp.get(date=d)
                        
                        tmp[str(d)] = e 
                    except:
                        tmp[str(d)] = ""
                
                res2[i.id] = tmp
        codes=Code.objects.all()
        context={'date':date,'instances':instances,'date_range':date_range,'codes':codes,'res':res,'today':date,'overwork':overworked_employes,'employe_on_mission':employe_on_mission,'res2':res2,'employe_to_disable':employe_to_disable,}
        return render(request,'pointage.html',context)


def logout_view(request):
    logout(request)
    return redirect(login_view)

def find_cell(date):
    if (date.day <=25):
        alphabet=chr(ord('B') + (date.day)%26 -1)
    else:
        alphabet="A"+chr(ord('A') + (date.day)%26)
    return (f"{alphabet}{date.month+13}")



def main_view(request,ID,year):
    i=Employe.objects.get(pk=ID)
    current_directory=os.getcwd()
    template_file = os.path.join(current_directory,"template.xlsx")
    output_file = os.path.join(current_directory,f"{request.user.username}{ID}.xlsx")
    current_directory = os.getcwd()

# Get the parent directory of the current working directory
    # Check if the output file exists
    file_path=os.path.join(current_directory,output_file)
    if os.path.exists(file_path):
        # If the output file exists, load it for editing
        os.remove(file_path)
        # If the output file doesn't exist, create a new workbook as a duplicate of the template
    copyfile(template_file, output_file)
    workbook = load_workbook(output_file)

    # Get the active worksheet (the first sheet by default)
    sheet = workbook.active

    # Add or modify data in the worksheet as needed
    sheet["AG6"].value = "" if f"{i.recruitment_date}" == 'None' else i.recruitment_date

# Set the value of the range "B8:N8" to i.function
    sheet["B8"].value = i.function

    # Set the value of the range "U6:X6" to i.id
    sheet["U6"].value = i.id

    # Set the value of the range "B6:N6" to i.name
    sheet["B6"].value = i.name

    # Set the value of the range "B7:N7" to i.last_name
    sheet["B7"].value = i.last_name

    # Set the value of the range "AG8:AJ8" to i.Date_Detach
    #sheet["AG8"].value = "" if f"{i.Date_Detach}" == 'None' else i.Date_Detach

    # Set the value of the range "B10:X10" to i.Adresse
    sheet["B10"].value = "" if f"{i.adresse}" == 'None' else i.adresse

    # Set the value of the range "AG9:AM9" to i.Affect_Origin
    #sheet["AG9"].value = "" if f"{i.Affect_Origin}" == 'None' else i.Affect_Origin

    # Set the value of the range "AG10:AJ10" to i.Situation_Familliale
    sheet["AG10"].value = "" if f"{i.familiy_situation}" == 'None' else i.familiy_situation

    # Set the value of the range "AG11:AH11" to i.Nbr_Enfants
    sheet["AG11"].value = "" if f"{i.numbre_of_children}" == 'None' else i.numbre_of_children

    # Set the value of the range "AO4:AT4" to the current year
    sheet["AO4"].value = year
    '''range=sheet['A118'].value
    sheet[range].value=request.user.profile.name
    column, row = coordinate_from_string(range)
    new_row = row + 1
    new_cell_ref = f"{column}{new_row}'''#this code for keeping logs for the user that enterd to pointage annee (needs date)
    try:
        code_list=Code_Employe.objects.filter(employe_id=ID,date__year=year)
        for code in code_list:
            base_row=ord('A')
            base_col=13
            date_to_treat=code.date
            base_col+=date_to_treat.month
            base_row+=date_to_treat.day
            if chr(base_row) >'Z':
                row='A'+chr(65 + (base_row - 65) % (90 - 65 + 1))
            else:
                row=chr(base_row)
            sheet[f'{row}{base_col}']=code.code_id

            
    except:

        pass

    # Save the workbook
    workbook.save(output_file)
    current_directory = os.getcwd()
    file_name = "convert.bat"
    file_path = os.path.join(current_directory, file_name)
    output_folder=current_directory #os.path.join(current_directory,"temp/")
    pdf_name=os.path.join(current_directory,f"{request.user.username}{ID}.pdf")
    excel_to_pdf(file_path,output_file,output_folder)
    pdf_file= pdf_name#os.path.join(output_folder,pdf_name)
    response = FileResponse(open(pdf_file, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = 'inline'
    create_log(request.user.username,"create",f"pointage annual {i}")
    return response

def login_view(request):
    #init_employe_pcpaye()
    #unit=Unite.objects.get(pk=1)
    #sync(unit)
    if request.user.is_authenticated:
        return redirect("menu_view")
    if request.method == 'POST':
        print(request)
        print(request.POST)
        unitname = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, unitename=unitname, password=password)
        if user is not None:
            login(request, user)
            create_log(user,"login"," ")
            return redirect('menu_view')  # Redirect to home page after successful login
        else:
            # Handle invalid login
            return render(request, 'login.html', {'error_message': 'Invalid username or password'})
    else:
        return render(request, 'login.html')






def menu_view(request):
    #init_sheet()
    
    if not request.user.is_authenticated:
        return redirect("login")
    id = request.user.profile.unite.id
    if request.user.profile.da == 1 and request.GET.get('unite'):
        id=request.GET.get('unite') 
    stations = Unite.objects.all()  #select * from unite
    unite = Unite.objects.get(id=id) #select * from unite where id==id
    yyyy = datetime.now().year
    mm = datetime.now().month
    if datetime.now().day < 16:
        mm = mm - 1
    if mm == 0 :
        mm = 12
        
    periode = f'{mm}{yyyy}'
    
    last_update =unite.last_update
    if 1 :
        employes=Employe.objects.filter(unite_id=id)
        MI=['1',"2","3","4","5","MS"]
        Conge=['C','CSS']
        for employe in employes:
            code_emp=''
            month_emp=''
            try:
                code_emps=Code_Employe.objects.filter(employe=employe,date__month=mm)
                if code_emps:
                    month_emp , _ = Month_stat.objects.get_or_create(employe=employe,period=periode)
                    for code_emp in code_emps:
                        if not code_emp.stored :
                            if code_emp.code_id == "T":
                                month_emp.travail+=1
                                month_emp.save()
                            try:
                                valid_date = ValidDate.objects.get(month=datetime.now().month)
                                if code_emp.date.day > valid_date.date_of_validation.day and code_emp.date.month == valid_date.month :
                                    if code_emp.code_id in MI:
                                        month_emp.mission+=1
                                    elif code_emp.code_id in "8":
                                        month_emp.absent+=1
                                    elif code_emp.code_id in Conge:
                                        month_emp.conge+=1
                                    elif code_emp.code_id  == "7":
                                        month_emp.abs_autorise+=1
                                    elif code_emp.code_id in "MLD":
                                        month_emp.mld+=1
                                    elif code_emp.code_id in "RS":
                                        month_emp.rs+=1
                                    elif code_emp.code_id in "6":
                                        month_emp.eve_fam+=1
                                    month_emp.save()
                            except:
                                pass
                                
                            code_emp.stored = True
                            code_emp.save()
                            log_action(request.user,ContentType.objects.get_for_model(Code_Employe),code_emp.id,str(code_emp),ADDITION,"to be saved")
            except Code_Employe.DoesNotExist:
                    continue
    
        
    return render(request, 'menu.html',{'id':id ,'unite':unite ,'stations':stations , 'today':datetime.now().day})


def add_employe(request,ID):
    #init_code()
    if request.user.profile.unite.id != ID and (request.user.profile.da != 1) :
        return redirect("menu_view")
    formset_diplomes=formset_factory(diplomesForm)
    formset_child=formset_factory(childForm)
    partnerform=partnerForm(prefix='partner')
    if request.user.profile.da == 1:
        if request.method == 'POST':
            form = EmployeFormForDg(request.POST,prefix='employe')
            diplomes=formset_diplomes(request.POST)
            children=formset_child(request.POST)
            partnerform=partnerForm(request.POST,prefix='partner')
            if form.is_valid():
                user=form.save()
                log_action(request.user,ContentType.objects.get_for_model(Employe),user.id,str(user),ADDITION,"to be saved")
                
                if partnerform.is_valid():
                    partner=partnerform.save(commit=False)
                    partner.id_employe=user
                    if partner.name:
                        partner.save()
                        log_action(request.user,ContentType.objects.get_for_model(Partner),partner.id,str(partner),ADDITION,"to be saved")
                if diplomes.is_valid():
                    for diplomeform in diplomes:
                        if diplomeform.cleaned_data:
                            diplome=diplomeform.save(commit=False)
                            diplome.id_employe=user
                            diplome.save()
                            log_action(request.user,ContentType.objects.get_for_model(Diplome),diplome.id,str(diplome),ADDITION,"to be saved")
                if children.is_valid():
                    for childform in children:
                        if childform.cleaned_data:
                            child=childform.save(commit=False)
                            child.id_employe=user
                            child.save()
                            log_action(request.user,ContentType.objects.get_for_model(Child),child.id,str(child),ADDITION,"to be saved")
                create_log(request.user.username,"add",f"employe {user}")
                return redirect('table_employe',ID)
        else :
            form = EmployeFormForDg(prefix="employe")
    else:
        if request.method == 'POST':
            form = EmployeForm(request.POST,prefix='employe')
            diplomes=formset_diplomes(request.POST)
            children=formset_child(request.POST)
            partnerform=partnerForm(request.POST,prefix="partner")
            if form.is_valid():
                user=form.save()
                user.unite=ID
                user.save()
                log_action(request.user,ContentType.objects.get_for_model(Employe),user.id,str(user),ADDITION,"to be saved")
                if partnerform.is_valid():
                    partner=partnerform.save(commit=False)
                    partner.employe_id=user.id
                    if partner.name:
                        partner.save()
                        log_action(request.user,ContentType.objects.get_for_model(Partner),partner.id,str(partner),ADDITION,"to be saved")
                if formset_diplomes.is_valid():
                    for diplomeform in diplomes:
                        if diplomeform.cleaned_data:
                            diplome=diplomeform.save(commit=False)
                            diplome.employe_id=user.id
                            diplome.save()
                            log_action(request.user,ContentType.objects.get_for_model(Code_Employe),diplome.id,str(diplome),ADDITION,"to be saved")
                if formset_child.is_valid():
                    for childform in children:
                        if childform.cleaned_data:
                            child=childform.save(commit=False)
                            child.employe_id=user.id
                            child.save()
                            log_action(request.user,ContentType.objects.get_for_model(Code_Employe),child.id,str(child),ADDITION,"to be saved")
                create_log(request.user.username,"add",f"employe {user}")
                return redirect('table_employe',ID)
        else:
            form = EmployeForm(prefix='employe')
            
    return render (request, 'add_employe.html', {'form': form,'id':ID,'diplomes':formset_diplomes,'children':formset_child,'partner':partnerform})

def table_employe(request,ID):
    if request.user.profile.unite.id != ID  and (request.user.profile.da != 1) :
        return redirect("menu_view")
    instances=Employe.objects.filter(unite=ID) #slecet * from employe where unite=ID
    return render(request,'table_employe.html',{'id':ID,'instances':instances,'da':request.user.profile.da})


        
def pointage_mois(request,ID):
    # if request.user.profile.unite.id != ID and (request.user.profile.da == 1) :
    #     return redirect("menu_view")
    instance=Employe.objects.get(pk=ID)
    current_year = datetime.now().year
    years = range(2023, current_year + 1) 
    return render(request,'mois_form.html',{'instance':instance,"years":years})


def affichage_mois(request,ID):
    
    month_names_french = {
        0: 'janvier',
        1: 'janvier',
        2: 'février',
        3: 'mars',
        4: 'avril',
        5: 'mai',
        6: 'juin',
        7: 'juillet',
        8: 'août',
        9: 'septembre',
        10: 'octobre',
        11: 'novembre',
        12: 'décembre',
    }
    month=int(request.POST.get("month"))
    if month == 0:
        given_year=request.POST.get('year')
        return redirect(main_view,ID,given_year)
    employe=Employe.objects.get(pk=ID)
    template_file = "template_mois.xlsx"
    output_file = f"{request.user.username}{ID}_mois.xlsx"
    current_directory = os.getcwd()


# Get the parent directory of the current working directory
    # Check if the output file exists
    template_file=os.path.join(current_directory,template_file)
    file_path=os.path.join(current_directory,output_file)
    if os.path.exists(file_path):
            # If the output file exists, load it for editing
            os.remove(file_path)
            # If the output file doesn't exist, create a new workbook as a duplicate of the template
    copyfile(template_file, output_file)
    workbook = load_workbook(output_file)

    # Get the active worksheet (the first sheet by default)
    sheet = workbook.active
    sheet["S4"]=f'16 {month_names_french[month]} 2024'
    sheet['W4']=f'15 {month_names_french[(int(month)+1)%12]} 2024'
    sheet['AC1']=f'{request.user.profile.unite}{datetime.now().date}'
    sheet['C11']=f'{employe.id}'
    sheet['H11']=f'{employe.name}'
    sheet['R11']=f'{employe.last_name}'
    sheet['R13']=f'{employe.function}'
    year=2024
    last_day=calendar.monthrange(year,month)[1]
    for rows in sheet['B17:AF17']:
        for cells in rows:
            cells.value=None
            cells.border=Border()
    for rows in sheet['B18:AF18']:
        for cells in rows:
            cells.value=None
            cells.border=Border() 
    start_date_obj = datetime.strptime(f'2024-{month}-16', "%Y-%m-%d")
    end_month=month+1
    if end_month==13:
        end_month=1
    end_date_obj = datetime.strptime(f'2024-{end_month}-15', "%Y-%m-%d")
    # Generate list of days as integers
    days_list = []
    current_date = start_date_obj
    while current_date <= end_date_obj:
        days_list.append(current_date.day)
        current_date += timedelta(days=1)
    row=17
    set_column='B'
    for counter in range(last_day):
        if set_column>'Z':
            column=f'A{chr(ord(set_column)%90+64)}'
        else:
            column=set_column
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        sheet[f'{column}{row}'].border=border_style
        sheet[f'{column}{row+1}'].border=border_style
        sheet[f'{column}{row}']=days_list[counter]
        if(days_list[counter]<16):
            date=datetime(year=2024,month=int(month)+1,day=days_list[counter])
        else:
            date=datetime(year=2024,month=month,day=days_list[counter])
        try:
            code=Code_Employe.objects.get(employe=employe,date=date)
            sheet[f'{column}{row+1}']=code.code_id
        except:
            pass
        
        set_column=chr(ord(set_column)+1)
    
    workbook.save(output_file)
    current_directory = os.getcwd()
    file_name = "convert.bat"
    file_path = os.path.join(current_directory, file_name)
    output_folder=current_directory #os.path.join(current_directory,"temp/")
    pdf_name=os.path.join(current_directory,f"{request.user.username}{ID}_mois.pdf")
    excel_to_pdf(file_path,output_file,output_folder)
    
    pdf_file=pdf_name #os.path.join(output_folder,pdf_name)
    response = FileResponse(open(pdf_file, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = 'inline'
    create_log(request.user.username,"create",f"pointage {month_names_french[month]} {employe}")
    return response




def pointage_mois_all(request,ID):
    # if request.user.profile.unite.id != ID and (request.user.profile.da == 1) :
    #     return redirect("menu_view")
    instance=Unite.objects.get(pk=ID)
    current_year = datetime.now().year
    years = range(2023, current_year + 1) 
    return render(request,'mois_form_all.html',{'instance':instance,"years":years})

def affichage_mois_all(request,ID):
    
    month_names_french = {
        0: 'janvier',
        1: 'janvier',
        2: 'février',
        3: 'mars',
        4: 'avril',
        5: 'mai',
        6: 'juin',
        7: 'juillet',
        8: 'août',
        9: 'septembre',
        10: 'octobre',
        11: 'novembre',
        12: 'décembre',
    }
    month=int(request.POST.get("month"))
    given_year=request.POST.get('year')
    if month == 0:
        
        return redirect(main_view_all,ID,given_year)
    employees=Employe.objects.filter(unite_id=ID,active=1)
    employees=employees.exclude(special=2)
    template_file = "template_mois.xlsx"
    output_file = f"{request.user.username}{ID}_unit_mois.xlsx"
    current_directory = os.getcwd()


# Get the parent directory of the current working directory
    # Check if the output file exists
    template_file=os.path.join(current_directory,template_file)
    file_path=os.path.join(current_directory,output_file)
    if os.path.exists(file_path):
        # If the output file exists, load it for editing
        os.remove(file_path)
        # If the output file doesn't exist, create a new workbook as a duplicate of the template
    copyfile(template_file, output_file)
    workbook=load_workbook(file_path)
    # Get the active worksheet (the first sheet by default)
    for employe in employees:
        sheet = workbook.copy_worksheet(workbook.active)
        sheet.title=f'{employe.id}'
        sheet["S4"]=f'16 {month_names_french[month]} 2024'
        sheet['W4']=f'15 {month_names_french[(int(month)+1)%12]} 2024'
        sheet['AC1']=f'{request.user.profile.unite}{datetime.now().date}'
        sheet['C11']=f'{employe.id}'
        sheet['H11']=f'{employe.name}'
        sheet['R11']=f'{employe.last_name}'
        sheet['R13']=f'{employe.function}'
        year=int(given_year)
        last_day=calendar.monthrange(year,month)[1]
        for rows in sheet['B17:AF17']:
            for cells in rows:
                cells.value=None
                cells.border=Border()
        for rows in sheet['B18:AF18']:
            for cells in rows:
                cells.value=None
                cells.border=Border() 
        start_date_obj = datetime.strptime(f'2024-{month}-16', "%Y-%m-%d")
        end_month=month+1
        if end_month==13:
            end_month=1
        end_date_obj = datetime.strptime(f'2024-{end_month}-15', "%Y-%m-%d")
        # Generate list of days as integers
        days_list = []
        current_date = start_date_obj
        while current_date <= end_date_obj:
            days_list.append(current_date.day)
            current_date += timedelta(days=1)
        row=17
        set_column='B'
        for counter in range(last_day):
            if set_column>'Z':
                column=f'A{chr(ord(set_column)%90+64)}'
            else:
                column=set_column
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            sheet[f'{column}{row}'].border=border_style
            sheet[f'{column}{row+1}'].border=border_style
            sheet[f'{column}{row}']=days_list[counter]
            if(days_list[counter]<16):
                date=datetime(year=2024,month=int(month)+1,day=days_list[counter])
            else:
                date=datetime(year=2024,month=month,day=days_list[counter])
            try:
                code=Code_Employe.objects.get(employe=employe,date=date)
                sheet[f'{column}{row+1}']=code.code_id
            except:
                pass
            
            set_column=chr(ord(set_column)+1)
    workbook.remove(workbook.active)    
    workbook.save(output_file)
    current_directory = os.getcwd()
    file_name = "convert.bat"
    file_path = os.path.join(current_directory, file_name)
    output_folder=current_directory #os.path.join(current_directory,"temp/")
    pdf_name=os.path.join(current_directory,f"{request.user.username}{ID}_unit_mois.pdf")
    excel_to_pdf(file_path,output_file,output_folder)
    
    pdf_file=pdf_name #os.path.join(output_folder,pdf_name)
    response = FileResponse(open(pdf_file, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = 'inline'
    create_log(request.user.username,"create",f"pointage {month_names_french[month]} {employe}")
    return response


def main_view_all(request,ID,year):
    employees=Employe.objects.filter(unite_id=ID,active=1)
    employees=employees.exclude(special=2)
    current_directory=os.getcwd()
    template_file = os.path.join(current_directory,"template.xlsx")
    output_file = os.path.join(current_directory,f"{request.user.username}{ID}_unit.xlsx")
    current_directory = os.getcwd()

# Get the parent directory of the current working directory
    # Check if the output file exists
    file_path=os.path.join(current_directory,output_file)
    if os.path.exists(file_path):
        # If the output file exists, load it for editing
        os.remove(file_path)
        # If the output file doesn't exist, create a new workbook as a duplicate of the template
    copyfile(template_file, output_file)
    workbook = load_workbook(output_file)
    for i in employees:
    # Get the active worksheet (the first sheet by default)
        sheet=workbook.copy_worksheet(workbook.active)
        sheet.title=f'{i.id}'

        # Add or modify data in the worksheet as needed
        sheet["AG6"].value = "" if f"{i.recruitment_date}" == 'None' else i.recruitment_date

    # Set the value of the range "B8:N8" to i.function
        sheet["B8"].value = i.function

        # Set the value of the range "U6:X6" to i.id
        sheet["U6"].value = i.id

        # Set the value of the range "B6:N6" to i.name
        sheet["B6"].value = i.name

        # Set the value of the range "B7:N7" to i.last_name
        sheet["B7"].value = i.last_name

        # Set the value of the range "AG8:AJ8" to i.Date_Detach
        #sheet["AG8"].value = "" if f"{i.Date_Detach}" == 'None' else i.Date_Detach

        # Set the value of the range "B10:X10" to i.Adresse
        sheet["B10"].value = "" if f"{i.adresse}" == 'None' else i.adresse

        # Set the value of the range "AG9:AM9" to i.Affect_Origin
        #sheet["AG9"].value = "" if f"{i.Affect_Origin}" == 'None' else i.Affect_Origin

        # Set the value of the range "AG10:AJ10" to i.Situation_Familliale
        sheet["AG10"].value = "" if f"{i.familiy_situation}" == 'None' else i.familiy_situation

        # Set the value of the range "AG11:AH11" to i.Nbr_Enfants
        sheet["AG11"].value = "" if f"{i.numbre_of_children}" == 'None' else i.numbre_of_children

        # Set the value of the range "AO4:AT4" to the current year
        sheet["AO4"].value = year
        '''range=sheet['A118'].value
        sheet[range].value=request.user.profile.name
        column, row = coordinate_from_string(range)
        new_row = row + 1
        new_cell_ref = f"{column}{new_row}'''#this code for keeping logs for the user that enterd to pointage annee (needs date)
        try:
            code_list=Code_Employe.objects.filter(employe_id=i.id,date__year=year)

            for code in code_list:
                base_row=ord('A')
                base_col=13
                date_to_treat=code.date
                base_col+=date_to_treat.month
                base_row+=date_to_treat.day
                if chr(base_row) >'Z':
                    row='A'+chr(65 + (base_row - 65) % (90 - 65 + 1))
                else:
                    row=chr(base_row)
                sheet[f'{row}{base_col}']=code.code_id

                
        except:

            pass

    # Save the workbook
    workbook.remove(workbook.active)
    workbook.save(output_file)
    current_directory = os.getcwd()
    file_name = "convert.bat"
    file_path = os.path.join(current_directory, file_name)
    output_folder=current_directory #os.path.join(current_directory,"temp/")
    pdf_name=os.path.join(current_directory,f"{request.user.username}{ID}_unit.pdf")
    err=excel_to_pdf(file_path,output_file,output_folder)
    print(err)
    pdf_file= pdf_name#os.path.join(output_folder,pdf_name)
    response = FileResponse(open(pdf_name, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = 'inline'
    create_log(request.user.username,"create",f"pointage annual {i}")
    return response



def update_employe(request,ID):
    employe=Employe.objects.get(id=ID)
    children = Child.objects.filter(id_employe=employe)
    diplomes = Diplome.objects.filter(id_employe=employe)
    try:
        partner=Partner.objects.get(id_employe=employe)
    except Partner.DoesNotExist:
        partner=None

    diplomes_setup=inlineformset_factory(Employe,Diplome,form=diplomesForm,extra=0,can_delete=True)
    child_setup=inlineformset_factory(Employe, Child,form=childForm,extra=0)
    
    if request.method == 'POST':
        if request.user.profile.da ==1:
            form= EmployeFormForDg(request.POST, instance=employe,prefix='employe')
        else:
            form = EmployeForm(request.POST, instance=employe,prefix='employe')
        formset_diplomes = diplomes_setup(request.POST, instance=employe, queryset=diplomes)
        formset_children = child_setup(request.POST, instance=employe, queryset=children)
        partner_form = partnerForm(request.POST,prefix='partner')
        
        if partner:
            partner_form = partnerForm(request.POST, instance=partner, prefix='partner')

        if form.is_valid():
            employe=form.save()
            if formset_diplomes.is_valid():
                for diplome in formset_diplomes:
                    if diplome["DELETE"].value():
                        d=diplome.instance
                        d.delete()
                    elif diplome.is_valid():
                        d=diplome.save(commit=False)
                        filled=any(diplome[field.name].value() for field in diplome if field.name not in [f'id_employe',f'DELETE'])
                        if filled:
                            d.id_employe=employe
                            d.save()
            if formset_children.is_valid():
                for child in formset_children:
                    if child["DELETE"].value():
                        c=child.instance
                        c.delete()
                    elif child.is_valid():
                        c=child.save(commit=False)
                        filled=any(child[field.name].value() for field in child if field.name not in [f'student',f'DELETE',f'id_employe'])
                        if filled:
                            c.id_employe=employe
                            c.save()
            
            if partner_form.is_valid():
                try:
                    request.POST["partner-DELETE"]
                    w=partner_form.instance
                    w.delete()
                except:
                    partner=partner_form.save(commit=False)
                    filled=any(field for field in partner_form if partner_form[field.name].value())
                    if filled:
                        partner.id_employe=employe
                        partner.save()
            return redirect(table_employe,employe.unite_id)
    else:
        if request.user.profile.da==1:
            form=EmployeFormForDg(instance=employe,prefix='employe')
        else:    
            form=EmployeForm(instance=employe,prefix='employe') 
        partnerform=partnerForm(prefix='partner')
        if partner:
            partnerform=partnerForm(instance=partner,prefix='partner')
            partnerform.fields['DELETE'] = forms.BooleanField(required=False, widget=forms.CheckboxInput(attrs={'class': 'delete-checkbox'}))
            
        formset_diplomes=diplomes_setup(instance=employe,queryset=diplomes)
        formset_children=child_setup(instance=employe,queryset=children)
        
    return render(request,'update_employe.html',{'i':employe,'form':form,'diplomes':formset_diplomes,'children':formset_children,'partner':partnerform})






def corriger_erreur(request):
    
    employees = Employe.objects.all()
    message=""
    if request.method == "POST":
        print(request.POST)
        date = request.POST.get("date")
        employe_id = request.POST.get("selected")
        code = Code_Employe.objects.filter(employe = employe_id , date=date)
        if code : 
            code[0].open_to_edit = True
            code[0].save()
            message = str(code[0])
        else :
            message = "NONE"
    return render(request , 'err.html' , {'employees':employees , 'message':message})

def remboursement_formule(request,ID):
    employe=Employe.objects.get(pk=ID)
    if request.method=='POST':
        form=remboursementForm(request.POST,instance=employe)
        if form.is_valid():
            form.save()
            return redirect(table_employe,employe.unite_id)
    else:
        form=remboursementForm(instance=employe)
        return render(request,'remboursement.html',{'employe':employe,'form':form})
    




def fiche_de_paie(validation_date , valid):
    month_names_french = {
        0: 'janvier',
        1: 'janvier',
        2: 'février',
        3: 'mars',
        4: 'avril',
        5: 'mai',
        6: 'juin',
        7: 'juillet',
        8: 'août',
        9: 'septembre',
        10: 'octobre',
        11: 'novembre',
        12: 'décembre',
    }
    end= {
    15: 'T',
    16: 'U',
    17: 'V',
    18: 'W',
    19: 'X',
    20: 'Y',
    21: 'Z',
    22: 'AA',
    23: 'AB',
    24: 'AC',
    25: 'AD',
    26: 'AE',
    27: 'AF',
    28: 'AG',
    29: 'AH',
    30: 'AI',
    31: 'AJ'
}
    
    endcell = end[validation_date.day]
    endtmp = None
    if len(endcell) >= 2:
        endtmp = endcell 
        endcell = 'Z'
    
    
    yyyy = datetime.now().year
    mm = datetime.now().month 
    current_directory=os.getcwd()
    
    if calendar.monthrange(yyyy, mm)[1] == 31 :
            default_excel =os.path.join(current_directory,"excel/default31.xlsx")
            final_cell = 'J'
    elif calendar.monthrange(yyyy, mm)[1] == 30 :
            default_excel = os.path.join(current_directory,"excel/default30.xlsx")
            final_cell = 'I'
    elif calendar.monthrange(yyyy, mm)[1] == 28 :
            default_excel = os.path.join(current_directory,"excel/default28.xlsx")
            final_cell = 'G'
    elif calendar.monthrange(yyyy, mm)[1] == 29 :
            default_excel = os.path.join(current_directory,"excel/default29.xlsx")
            final_cell = 'H'
    
            
    
    employees = Employe.objects.filter(active=1)
    if not valid:
        path = os.path.join(current_directory,f'excel/fiche_de_pointage{month_names_french[mm]}{yyyy}NonValide.xlsx')
    else :
        path = os.path.join(current_directory,f'excel/fiche_de_pointage{month_names_french[mm]}{yyyy}.xlsx')

    shutil.copy(default_excel,path)
    worbook = load_workbook(path)
    sheet = worbook.active 
    i = 20 
    sheet[f'F6'] = f'{date(year=yyyy,month=mm,day=1)}'
    month_periode = datetime.now().month - 1
    if month_periode == 0 :
        month_periode = 12
    tt =92
    t = 92
    for employe in employees:
        mission = {
        "5":0 , "4":0 , "3":0 , "2" : 0 , "1" : 0 , "MS":0 
        }   
        tmp_date = datetime.now().replace(day=1)
        sheet[f'C{i}'] = f'{employe.name} {employe.last_name}'
        sheet[f'B{i}'] = employe.id
        sheet[f'E{i}'] = employe.function
        if employe.special==2:
            for cell in range(ord('F'),ord(endcell)+1):
                sheet[f'{chr(cell)}{i}']='T'
            if endtmp is not None:
                c = endtmp[1]
                for cell in range(ord('A'),ord(c)+1):
                    sheet[f'A{chr(cell)}{i}']='T'
        else:
            if employe.refund_by_month != 0 :
                sheet[f'D{i}'] = f'{employe.refund_by_month}'
                employe.refund_total -= employe.refund_by_month
                employe.refund_by_month = 0 
                employe.save()
            for cell in range(ord('F'),ord(endcell)+1):
                code = Code_Employe.objects.filter(employe=employe,date=tmp_date)
                try:
                    if str(code[0]) in mission:
                        sheet[f'{chr(cell)}{i}'] = "MI"
                        mission[str(code[0])] += 1
                    else:
                        sheet[f'{chr(cell)}{i}'] = str(code[0].code)
                except:
                    sheet[f'{chr(cell)}{i}'] = " "
                tmp_date = tmp_date + timedelta(days=1)
            if endtmp is None:
                if endcell != 'Z':
                    for cell in range(ord(endcell)+1 , ord('Z') + 1): 
                        sheet[f'{chr(cell)}{i}'] = "T"
                for cell in range(ord('A'),ord(final_cell)+1):
                    sheet[f'A{chr(cell)}{i}'] = "T"
            else:
                c = endtmp[1]
                for cell in range(ord('A'),ord(c)+1):
                    code = Code_Employe.objects.filter(employe=employe,date=tmp_date)
                    try:
                        print(code[0])
                        if str(code[0]) in mission:
                            sheet[f'{chr(cell)}{i}'] = "MI"
                            mission[str(code[0])] += 1
                        else:
                            sheet[f'A{chr(cell)}{i}'] = str(code[0].code)
                    except:
                        sheet[f'A{chr(cell)}{i}'] = " " 
                    tmp_date = tmp_date + timedelta(days=1)  
                for cell in range(ord(c),ord(final_cell)+1):
                    sheet[f'A{chr(cell)}{i}'] = "T"
            
            month_stat=None
            
            try :
                month_stat = Month_stat.objects.get(employe=employe , period = f'{month_periode}{yyyy}')
                # print("have : ", month_stat)
                sheet[f'AL{i}'].value += f'+{month_stat.conge}' 
                sheet[f'AM{i}'].value += f'+{month_stat.absent}'
                sheet[f'AN{i}'].value += f'+{month_stat.abs_autorise}'
                sheet[f'AO{i}'].value += f'+{month_stat.rs}'
                sheet[f'AP{i}'].value += f'+{month_stat.eve_fam}'
                sheet[f'AQ{i}'].value += f'+{month_stat.mission}'
                sheet[f'AR{i}'].value += f'+{month_stat.mld}'
            except:
                pass
            
            
            for k,v in mission.items() :
                if v != 0 :
                    sheet[f'F{tt}'].value = f'{employe.name} {employe.last_name} : {v} j mission en code ({k})'
                    tt += 1
            
            if month_stat:
                if month_stat.absent != 0:
                    # print("Employe name : " , employe.name)
                    sheet[f'AC{t}'].value = f'{employe.name} {employe.last_name} : {month_stat.absent} j Absence  sur le mois {month_names_french[mm-1]}'
                    t = t+1
                if month_stat.abs_autorise != 0:
                    sheet[f'AC{t}'].value = f'{employe.name} {employe.last_name} : {month_stat.abs_autorise} j Absence  Autorise sur le mois {month_names_french[mm-1]}'
                    t = t+1
                if month_stat.conge != 0:
                    sheet[f'AC{t}'].value = f'{employe.name} {employe.last_name} : {month_stat.conge} j Conge {month_names_french[mm-1]}'
                    t = t+1
                if month_stat.rs != 0:
                    sheet[f'AC{t}'].value = f'{employe.name} {employe.last_name} : {month_stat.rs} j Recuperation {month_names_french[mm-1]}'
                    t = t+1
                if month_stat.mld != 0:
                    sheet[f'AC{t}'].value = f'{employe.name} {employe.last_name} : {month_stat.mld} j Maladie {month_names_french[mm-1]}'
                    t = t+1
                
            i = i + 1
    zs=Employe.objects.filter(active=2)
    for z in zs:
        sheet[f'C{i}'] = f'{z.name} {z.last_name}'
        sheet[f'B{i}'] = z.id
        sheet[f'E{i}'] = z.function
        z.actif=3
        z.save()
        sheet.merge_cells(f'F{i}:AG{i}')
        sheet[f'F{i}']='STC'
    if valid:
        sheet['F14'].value = f'validee le {validation_date} '
    worbook.save(f"{path}")
    
    
    
    
    
    return redirect('menu_view')

# def download_excel(request,name):
#     path = f"/home/petromag-dz/public_html/pointage/excel/{name}.xlsx"
#     print("DOWNLOAD")
#     if os.path.exists(path):
#         response = FileResponse(open(path, 'rb'))
#         response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(name)
#         return response
#     else:
#         # Handle the case where the file is not found
#         return HttpResponse("The file you are trying to download does not exist.", status=404)
def desactiver(request,ID):
    emp = Employe.objects.get(pk=ID)
    
    emp.active=2
    emp.save()
    
    
    return redirect('menu_view')

def activer(request,ID):
    emp = Employe.objects.get(pk=ID)
    
    emp.active=1
    emp.save()
    
    
    return redirect('menu_view')

def validation_fdp(request):
    month_names_french = {
        0: 'janvier',
        1: 'janvier',
        2: 'février',
        3: 'mars',
        4: 'avril',
        5: 'mai',
        6: 'juin',
        7: 'juillet',
        8: 'août',
        9: 'septembre',
        10: 'octobre',
        11: 'novembre',
        12: 'décembre',
    }
    if request.user.profile.da != 1:
        return redirect("menu_view")
    employe_rembourse = Employe.objects.filter(refund_total__gt=0)
    e = ValidDate.objects.filter(month=datetime.now().month).exists()
    if request.method == "POST":
        validation_date = datetime.strptime(request.POST.get("date"), '%Y-%m-%d')
        for emp in employe_rembourse:
            value = request.POST.get(f'{emp.id}Rmois')
            emp.refund_by_month = int(value) 
            emp.save()
        if 'button_without_validation' in request.POST:
            valid = False
        elif 'button_with_validation' in request.POST:
            valid = True
            validation_date = datetime.strptime(request.POST.get("date"), '%Y-%m-%d')
            valid_date = ValidDate.objects.create(
                                            date_of_validation=validation_date ,
                                            nom = "p" ,
                                            month=datetime.now().month
                                            ) 
            valid_date.save()
        fiche_de_paie(validation_date,valid)
        
    
    
    p = f'{month_names_french[datetime.now().month]}{datetime.now().year}'   
    pp = f'{p}NonValide'
    return render(request , 'validation_fdp.html', {"today":datetime.now().day, 'employees': employe_rembourse  ,
                                                    "exist":e , "p":p , "pp":pp})



def download_excel(request, p):
    current_directory=os.getcwd()
    path = f"{current_directory}/excel/fiche_de_pointage{p}.xlsx"
    if os.path.exists(path):
        response = FileResponse(open(path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="fiche_de_pointage{p}.xlsx"'
        return response
    else:
        # Handle the case where the file is not found
        return HttpResponse("The file you are trying to download does not exist.", status=404)


def add_profile(request):
    #init_code()
    if request.user.profile.da != 1:
        return redirect("menu_view")
    stations = Unite.objects.all()
    employe = Employe.objects.all()
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        unite_id = request.POST.get('unite')
        da = request.POST.get('da')
        user = User.objects.create_user(username=username,password=password)
        Profile.objects.create(user=user, unite_id=unite_id, da=da)
        return redirect('login')  # Redirect to login page after user creation
    
    return render(request, 'add_profile.html',{'stations':stations,'instances':employe})

def synthese(request,ID):
    employe=Employe.objects.get(pk=ID)
    template_file = "template_synthese.xlsx"
    output_file = f"{request.user.username}{ID}_synthese.xlsx"
    current_directory = os.getcwd()


# Get the parent directory of the current working directory
    # Check if the output file exists
    template_file=os.path.join(current_directory,template_file)
    file_path=os.path.join(current_directory,output_file)
    if os.path.exists(file_path):
        # If the output file exists, load it for editing
        workbook = load_workbook(output_file)
    else:
        # If the output file doesn't exist, create a new workbook as a duplicate of the template
        copyfile(template_file, output_file)
        workbook = load_workbook(output_file)

    # Get the active worksheet (the first sheet by default)
    sheet = workbook.active
    sheet["B7"]=employe.name
    sheet["F7"]=employe.last_name
    sheet["B9"]=employe.date_of_birth
    sheet["D9"]=employe.place_of_birth
    sheet["H9"]=employe.wilaya_of_birth
    sheet["B11"]=employe.adresse
    sheet["H11"]=employe.adresse_wilaya
    sheet["B13"]=employe.father_name
    sheet["E13"]=employe.mother_name
    sheet["I13"]=employe.phone
    if employe.familiy_situation:
        sheet["B15"]="Marié"
        sheet["D15"]=employe.numbre_of_children
    else :
        sheet["B15"]="Célibataire"
    sheet["F15"]=employe.blood_type
    sheet["H15"]=employe.cnas_number
    sheet["B17"]=employe.function
    sheet["G17"]=employe.position
    sheet["I17"]=employe.enterprise
    sheet["B19"]=employe.recruitment_date
    sheet["E19"]=employe.department
    sheet["H19"]=employe.service
    sheet["B21"]=employe.contract_number
    sheet["D21"]=employe.contract_effective_date
    sheet["F21"]=employe.contract_validation_date
    sheet["H21"]=employe.contract_termination_date
    sheet["C23"]=employe.national_service_departure_date
    sheet["E23"]=employe.national_service_returne_date
    sheet["G23"]=employe.national_service_recall_departure_date
    sheet["I23"]=employe.national_service_recallt_return_date
    sheet["D25"]=employe.account_number
    sheet["G25"]=employe.account_key
    sheet["I25"]=employe.account_agency
    sheet["C27"]=employe.driver_license_number
    sheet["E27"]=employe.driver_license_established_date
    sheet["G27"]=employe.driver_license_experation_date
    sheet["I27"]=employe.driver_license_type
    sheet["B29"]=employe.cni_number
    sheet["E29"]=employe.cni_established_date
    sheet["G29"]=employe.cni_established_by
    diplomes=Diplome.objects.filter(id_employe=employe)
    row=33
    if diplomes and len(diplomes)<=4:
        for diplome in diplomes:
            sheet[f'A{row}']=diplome.establishment
            sheet[f'D{row}']=diplome.entry_date
            sheet[f'E{row}']=diplome.end_date
            sheet[f'F{row}']=diplome.diplome_name
            row+=1
    else:
        for i in range(row,row+5):
            sheet[f'A{i}']=""
            sheet[f'D{i}']=""
            sheet[f'E{i}']=""
            sheet[f'F{i}']=""
    try:
        partner=Partner.objects.get(id_employe=employe)
        sheet['B42']=partner.name
        sheet['F42']=partner.last_name
        sheet['B44']=partner.date_of_birth
        sheet['F44']=partner.place_of_birth
        sheet['I44']=partner.wilaya_of_birth
        sheet['C46']=partner.marriage_date
        sheet['F46']=partner.partner_salary
        if not partner.partner_salary:
            sheet['I46']='yes'
    except:
        sheet['B42']=""
        sheet['F42']=""
        sheet['B44']=""
        sheet['F44']=""
        sheet['I44']=""
        sheet['C46']=""
        sheet['F46']=""
        sheet['I46']=""
    children=Child.objects.filter(id_employe=employe)
    row=51
    if children and len(children)<=4:
        for child in children:
            sheet[f'A{row}']=child.name
            sheet[f'B{row}']=child.last_name
            sheet[f'D{row}']=child.date_of_birth
            sheet[f'F{row}']=child.place_of_birth
            sheet[f'H{row}']=child.student
            sheet[f'I{row}']=child.af
            row+=1
    else:
        for i in range(row,row+5):
            sheet[f'A{i}']=""
            sheet[f'B{i}']=""
            sheet[f'D{i}']=""
            sheet[f'F{i}']=""
            sheet[f'H{i}']=""
            sheet[f'I{i}']=""


    workbook.save(output_file)
    current_directory = os.getcwd()
    file_name = "convert.bat"
    file_path = os.path.join(current_directory, file_name)
    output_folder=current_directory #os.path.join(current_directory,"temp/")
    pdf_name=os.path.join(current_directory,f"{request.user.username}{ID}_synthese.pdf")
    excel_to_pdf(file_path,output_file,output_folder)
    
    pdf_file=pdf_name #os.path.join(output_folder,pdf_name)
    response = FileResponse(open(pdf_file, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = 'inline'
    return response

def mission(request,ID):
    employe=Employe.objects.get(id=ID)
    unites=Unite.objects.all()
    if request.method=='GET':
        return render(request,'mission.html',{'employe':employe,'unites':unites})
    else:
        mission=Mission()
        date_str=request.POST.get('start_date')
        date = datetime.strptime(date_str, '%Y-%m-%d')
        check_date=Mission.objects.filter(start_date=date,employe_id=ID)
        if (check_date):
            message='employe already have a mission on this date'
            messages.error(request,message)
            return render(request,'mission.html',{'employe':employe,'unites':unites})
        if date < datetime.today():
            message='invalid date value '
            messages.error(request,message)
            return render(request,'mission.html',{'employe':employe,'unites':unites})
        unite_id=request.POST.get('unite')

        unite=Unite.objects.get(pk=unite_id)
        if unite.id==employe.unite_id:
            message='invalid unite please try again'
            messages.error(request,message)
            return render(request,'mission.html',{'employe':employe,'unites':unites})
        unite_2_id=request.POST.get('unite_1')
        if unite_2_id:
            all_unites=unite_id +' '+ unite_2_id
            validation_needed=2
            unite_3_id=request.POST.get('unite_2')
            if unite_3_id:
                all_unites=all_unites +' '+ unite_3_id
                validation_needed=3
        else:
            all_unites=unite_id
            validation_needed=1
        mission.start_date=date
        mission.employe=employe
        mission.total_validation=validation_needed
        mission.current_unite=unite
        mission.unites=all_unites
        mission.save()
        create_log(request.user.username,"create",f"mission {mission}")
        return redirect(table_employe,employe.unite_id)
    
def serialize_instance(instance):
    data = {}
    for field in instance._meta.fields:
        value = getattr(instance, field.name)
        if isinstance(value, (datetime, date)):
            value = value.isoformat()
        elif isinstance(field, models.ForeignKey):
            related_model = field.related_model

            related_instance = getattr(instance, field.name)
            if related_instance:
                value = f"{related_model._meta.model_name}:{related_instance.id}"
            else:
                value = None
        data[field.name] = value
    return data

from django.contrib.admin.models import LogEntry
from django.apps import apps
import requests
def sync(unit):
    if check_web_server():
        print ('server is up')
        try:
            log_enteries=LogEntry.objects.filter(action_time__date__gt=unit.last_update)
            instances_data=[]
            for log in log_enteries:
                model_class=log.content_type.model_class()
                if model_class.__name__ != 'Profile' and log.action_flag==1:
                    print(model_class,log.object_id)
                    instance=model_class.objects.get(id=log.object_id)
                    instance_data={
                        "model":log.content_type.model,
                        "instance_id":instance.id,
                        "data": serialize_instance(instance)
                    }
                    instances_data.append(instance_data)
            response=requests.post("http://127.0.0.1:9000/fetch/",json=instances_data)
            if response.status_code==200:
                response_data=response.json()
                update=response_data['update']
                for header,id in update.items():
                    model,old_id=header.split('|')
                    model_class=apps.get_model(app_label='pointage',model_name=model)
                    old_instance=model_class.objects.get(pk=old_id)
                    old_instance.id=id
                    old_instance.save()
                unit.last_update=datetime.now().date()
                unit.save()
                return 1
            else:
                return 0
        except Exception as e:
            print(e)
            return -1
    else:
        print('server is down')
        return -1    



from django.contrib.admin.models import LogEntry
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone

def log_action(user, content_type, object_id, object_repr, action_flag, change_message=None):
    """
    Function to create a log entry in Django's admin_log table.
    
    Parameters:
    - user: User instance who performed the action.
    - content_type: ContentType instance of the affected object.
    - object_id: ID of the affected object.
    - object_repr: String representation of the affected object.
    - action_flag: Action flag (ADDITION, CHANGE, DELETION).
    - change_message: Optional message describing the change.
    """
    LogEntry.objects.log_action(
        user_id=user.pk,
        content_type_id=content_type.pk,
        object_id=object_id,
        object_repr=object_repr,
        action_flag=action_flag,
        change_message=change_message,
    )

def check_web_server():
    url = 'http://127.0.0.1:9000/'

    try:
        response = requests.head(url)
        if response.status_code == 200:
            return 1
        else:
            return 0
    except requests.ConnectionError:
        return 0