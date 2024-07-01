#the first part is a script that takes information from a sheet and create instances for each employe in that sheet If you want to use it you should modify model so that it won't check for pk
from django.db import connection
from .models import *
from openpyxl import load_workbook
from datetime import date, datetime 
import glob
import os
import subprocess
def init_code_annual():
    Code_Employe.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='pointage_code_employe';")

# Step 3: Vacuum the database to reclaim storage space
    with connection.cursor() as cursor:
        cursor.execute("VACUUM;")
    current_dir = os.getcwd()
    parent_dir=os.path.join(current_dir,"test")
# Search for Excel files in the parent directory
    excel_files = glob.glob(os.path.join(parent_dir, '*.xlsx'))
    column=['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']
    rows=14
    for file in excel_files:
        workbook=load_workbook(file)
        sheet_names = workbook.sheetnames
    # Print the list of sheet names
        for name in sheet_names:
            worksheet=workbook[name]
            emp=Employe.objects.get(pk=int(name))
            for row in range(rows,26):
                
                month=row-13
                year=2024
                for col in column:
                    try:
                        code_emp=Code_Employe()
                        code_emp.employe=emp
                        day=column.index(col)+1
                        code=Code.objects.get(pk=worksheet[f"{col}{row}"].value)
                        code_emp.code=code
                        code_emp.date=date(year,month,day)
                        code_emp.save()
                    except Exception as e:
                        #print(col,row,emp.id,worksheet[f"{col}{row}"].value)
                        
                        pass


def init_sheet():
    Employe.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='pointage_employe';")

# Step 3: Vacuum the database to reclaim storage space
    with connection.cursor() as cursor:
        cursor.execute("VACUUM;")
    
    file=r'C:\Users\lenovo\OneDrive\Bureau\git_repo\pointage_for_windows\pointage\lISTE DU PERSONNEL.xlsx'
    workbook=load_workbook(file)
    sheet_names=workbook.sheetnames
    i=0
    column=['C','D','E','F']
    for sheet in sheet_names:
        worksheet=workbook[sheet]
        row=14
        i+=1
        while worksheet[f'{column[0]}{row}'].value:
            employe=Employe()
            employe.id=worksheet[f'{column[0]}{row}'].value
            employe.name=worksheet[f'{column[1]}{row}'].value
            employe.last_name=worksheet[f'{column[2]}{row}'].value
            employe.function=worksheet[f'{column[3]}{row}'].value
            employe.unite_id=i
            row+=1
            employe.save()

def init_code():
    Code_Employe.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='pointage_code_employe';")

# Step 3: Vacuum the database to reclaim storage space
    with connection.cursor() as cursor:
        cursor.execute("VACUUM;")
    current_dir = os.getcwd()
    parent_dir=os.path.join(current_dir,"test")
# Search for Excel files in the parent directory
    excel_files = glob.glob(os.path.join(parent_dir, '*.xlsx'))
    for file in excel_files:
        workbook=load_workbook(file)
        sheet_names = workbook.sheetnames

    # Print the list of sheet names
        for name in sheet_names:
            worksheet=workbook[name]
            try:
                emp=Employe.objects.get(pk=int(name))
            except:
                continue
            days=[]
            codes=[]
            date_current = worksheet['S4'].value
            year_of_the_file=date_current.year
            month_of_the_file=date_current.month
            try:
                if worksheet['B17'].value.isdigit():
                    index=17
                else:
                    index=18
            except:
                index=17

            for row in worksheet[f'B{index}:AF{index}']:
                for cell in row:
                    days.append(cell.value)
            for row in worksheet[f'B{index+1}:AF{index+1}']:
                for cell in row:
                    codes.append(cell.value)

            #if file=='C:\Users\lenovo\OneDrive\Bureau\pointage\test\Copie de FEUILLE DE POSITION ALGER JUIN 2024 (2)' and :
            for iterator in range(len(days)):
                
                code_emp=Code_Employe()
                code_iter=codes[iterator]
                try:
                    code_emp.code=Code.objects.get(pk=codes[iterator])
                except:
                    if codes[iterator] :
                        print(f'wrong code {codes[iterator]} in {file} for {emp.id} at {iterator}')
                    continue
                if int(iterator)>=16:
                    if date_current.month==12:
                        code_emp.date=date_current.replace(day=int(days[iterator]),month=1,year=year_of_the_file+1)
                    else:    
                        code_emp.date=date_current.replace(day=int(days[iterator]),month=date_current.month+1)
                else:
                    code_emp.date=date_current.replace(day=int(days[iterator]))
                code_emp.employe=emp
                code_emp.save()

def init_employe():
    current_dir = os.getcwd()
    parent_dir=os.path.join(current_dir,"creer")
# Search for Excel files in the parent directory
    excel_files = glob.glob(os.path.join(parent_dir, '*.xlsx'))
    if excel_files:
        for file in excel_files:
            workbook=load_workbook(file)
            sheet_names=workbook.sheetnames
            for sheet_name in sheet_names:
                worksheet=workbook[sheet_name]
                if worksheet['B7'].value:
                    try:
                        employe=Employe.objects.get(name=worksheet['B7'].value ,last_name=worksheet['F7'].value)
                    except:
                        employe=Employe()
                    employe.name=worksheet['B7'].value
                    employe.last_name=worksheet['F7'].value
                    employe.date_of_birth=worksheet['B9'].value
                    employe.place_of_birth=worksheet['D9'].value
                    employe.wilaya_of_birth=worksheet['H9'].value
                    employe.adresse=worksheet['B11'].value
                    employe.adresse_wilaya=worksheet['H11'].value
                    employe.father_name=worksheet['B13'].value
                    employe.mother_name=worksheet['E13'].value
                    employe.phone=worksheet['I13'].value
                    employe.familiy_situation=worksheet['B15'].value
                    employe.numbre_of_children=worksheet['D15'].value
                    employe.blood_type=worksheet['F15'].value
                    employe.cnas_number=worksheet['H15'].value
                    employe.function=worksheet['B17'].value
                    employe.position=worksheet['G17'].value
                    if not employe.unite:
                        unite=["ALGER","MECHRIA","TAMANRASSET","AIN SEFRA","IN SALAH"]
                        index=unite.index(employe.position)
                        employe.unite=Unite.objects.get(id=index+1)

                    employe.enterprise=worksheet['I17'].value
                    employe.recruitment_date=worksheet['B19'].value
                    employe.department=worksheet['E19'].value
                    employe.service=worksheet['H19'].value
                    employe.contract_number=worksheet['B21'].value
                    employe.contract_effective_date=worksheet['D21'].value
                    employe.contract_validation_date=worksheet['F21'].value
                    employe.contract_termination_date=worksheet['H21'].value
                    if worksheet['C23'].value != 'EXEMPLTE':
                        employe.national_service_departure_date=worksheet['C23'].value
                        employe.national_service_returne_date=worksheet['E23'].value
                        employe.national_service_recall_departure_date=worksheet['G23'].value
                        employe.national_service_recallt_return_date=worksheet['I23'].value
                    employe.account_number=worksheet['D25'].value
                    employe.account_key=worksheet['G25'].value
                    employe.account_agency=worksheet['I25'].value
                    if worksheet['C27'].value:
                        employe.driver_license_number=worksheet['C27'].value
                        employe.driver_license_established_date=worksheet['E27'].value
                        employe.driver_license_experation_date=date_handling(worksheet['G27'].value)
                        employe.driver_license_type=worksheet['I27'].value
                    employe.cni_number=worksheet['B29'].value
                    employe.cni_established_date=worksheet['E29'].value
                    employe.cni_established_by=worksheet['G29'].value
                    employe.save()
                    cell=worksheet['A33'].value
                    row=33
                    while cell:
                        diplome=Diplome()
                        diplome.establishment=cell
                        if worksheet[f'D{row}'].value:
                            diplome.entry_date=worksheet[f'D{row}'].value
                        diplome.end_date=date_handling(worksheet[f'E{row}'].value)
                        diplome.diplome_name=worksheet[f'F{row}'].value
                        diplome.id_employe=employe
                        diplome.save()
                        row+=1
                        cell=worksheet[f'A{row}'].value
                    if worksheet['B42'].value:
                        partner=Partner()
                        partner.id_employe=employe
                        partner.name=worksheet['B42'].value
                        partner.last_name=worksheet['F42'].value
                        partner.date_of_birth=worksheet['B44'].value
                        partner.place_of_birth=worksheet['F44'].value
                        partner.wilaya_of_birth=worksheet['I44'].value
                        partner.marriage_date=date_handling(worksheet['C46'].value)
                        partner.partner_salary=worksheet['F46'].value
                        partner.save()
                    cell=worksheet['A51'].value
                    row=51
                    while cell:
                        child=Child()
                        child.id_employe=employe
                        child.name=worksheet[f'A{row}'].value
                        child.last_name=worksheet[f'B{row}'].value
                        
                        child.date_of_birth=worksheet[f'D{row}'].value
                        student=worksheet['F51'].value
                        if student:
                            if student == 'NON':
                                child.student=False
                            else:
                                child.student=True
                        child.af=worksheet[f'I{row}'].value
                        child.save()
                        row+=1
                        cell=worksheet[f'A{row}'].value

def excel_to_pdf(file_path,*args):
    try:
        # Construct the command to execute the batch script
        command = [file_path] + list(args)
        print(command)
        
        # Execute the command
        process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
        stdout, stderr = process.communicate()

        # Check if the process exited successfully
        if process.returncode == 0:
            # Process output
            output = stdout.decode('utf-8')
            return output
        else:
            # Error occurred
            error = stderr.decode('utf-8')
            return f"Err: {error}"
    except Exception as e:
        return f"Error: {str(e)}"
    
def date_handling(input):
    
    if isinstance(input, datetime):
        output= input.date()  # Convert to date
    elif input:
        try:
                # Assuming date is in 'MM/DD/YYYY' format
            output = datetime.strptime(str(input), '%m/%d/%Y').date()
        except ValueError:
            output= None  # Handle invalid date formats
    else :
        output=None
    return output


def codage(table,info,operation):
    current_directory=os.getcwd() #
    filename='changes.txt'
    file_path=os.path.join(current_directory,filename)
    try:
        with open (file_path,'a') as file3:
            string = str(table)+'||'+str(operation)+'||'
            file3.write (string)
            if table == 68:#employe
                field_names = [field.name for field in Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many ]
            else :
                field_names = [field.name for field in Code_Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many]
            
            data={field : getattr(info,field) for field in field_names}
            for key in data.keys():
                file3.write (str(data[key])+'||')
                
            file3.write (str('\n'))
        return 1
    except :
        return 0
    
def decodage():

    current_directory=os.getcwd() #
    filename='changes.txt'
    file_path=os.path.join(current_directory,filename)
    try:
        
        with open(file_path,'r') as file3:
            for line in file3:
                list_elements=line.split('||')
                if  list_elements[0]==str(68):#employe
                    field_names = [field.name for field in Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many ]
                else :
                    field_names = [field.name for field in Code_Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many]
                del list_elements[0]
                operation=list_elements.pop(0)
                if operation ==str(1):
                    code_emp=Code_Employe()
    except:
        pass
    return 1

def create_log(profile,operation,table):
    new_log=History()
    new_log.user=profile
    new_log.operation=operation
    new_log.table=table
    new_log.date=datetime.now()
    new_log.save()

def init_employe_pcpaye():
    data = {
        "T": "Travail sur site",
        "RS": "Recuperation systeme",
        "FDS": "Fin de semaine avec salaire sans et panier",
        "M": "Arret de travail pour maladie",
        "6": "Conge special -evenement familiale- (naissance, mariage, deces,.....)",
        "7": "Absence autorisee sans salaire et sans IZCV",
        "8": "Absence non autorisee sans salaire et sans IZCV",
        "9": "Mise a pied",
        "C": "Conge paye (deduit du CP)",
        "CSS": "Conge sans solde",
        "7D": "Mise en disponibilite sans salaire et sans IZCV",
        "R1": "Recuperation sans IZCV deduit du reliquat",
        "JF": "Jour ferie avec salaire et sans panier",
        "A": "Arret de travail pour accident de travail",
        "RDC": "Rayer du contrat (deces, demission, ......)",
        "MLD": "Maladie longue duree (declare a la CNAS)",
        "V": "Delai de route",
        "X": "Abattement code (6)",
        "1": "Mission sans IZCV et sans reliquat",
        "2": "Mission autre Zone",
        "3": "Mission avec IZCV et reliquat (mission du nord vers le sud)",
        "4": "Mission sans IZCV avec CR sans IZCV (declenche depuis le domicile pendent le CR)",
        "5": "Mission avec IZCV et CR sans IZCV",
        "MS": "Mission sans IZCV avec reliquat des FDS et Ferie",
        "CR": "Congé de récupération",
        "CP": "Congé payé",
    }

    # Create instances
    for code_id, description in data.items():
        code_instance, created = Code.objects.get_or_create(id=code_id, defaults={'description': description})
        if created:
            print(f"Created Code: {code_id} - {description}")
        else:
            print(f"Code already exists: {code_id}")

    Employe.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='pointage_employe';")

# Step 3: Vacuum the database to reclaim storage space
    with connection.cursor() as cursor:
        cursor.execute("VACUUM;")

    current_directory=os.getcwd()
    pcpaye_directory=os.path.join(current_directory,'pcpaye')
    
    excel_files=glob.glob(os.path.join(pcpaye_directory,'*.xlsx'))
    if excel_files:
        for file in excel_files:
            workbook=load_workbook(file)
            sheet_names=workbook.sheetnames
            for sheet_name in sheet_names:
                row=11
                worksheet=workbook[sheet_name]
                while worksheet[f"B{row}"].value:
                    employe,created=Employe.objects.get_or_create(pk=int(worksheet[f"B{row}"].value))
                    if created:
                        employe.name=worksheet[f"C{row}"].value
                        employe.last_name=worksheet[f"D{row}"].value
                        employe.date_of_birth=date_handling(worksheet[f"E{row}"].value)
                        employe.place_of_birth=worksheet[f"F{row}"].value
                        employe.adresse=worksheet[f"H{row}"].value
                        employe.father_name=worksheet[f"I{row}"].value
                        employe.mother_name=str(worksheet[f"J{row}"].value +" "+ worksheet[f"K{row}"].value)
                        if worksheet[f"L{row}"].value == 'M':
                            employe.familiy_situation=1
                            employe.numbre_of_children=int(worksheet[f"M{row}"].value[2])
                        else:
                            employe.familiy_situation=0
                        employe.recruitment_date=date_handling(worksheet[f"O{row}"].value)
                        employe.cnas_number=int(worksheet[f"P{row}"].value)
                        employe.phone=[int(worksheet[f"Q{row}"].value) if worksheet[f"Q{row}"].value else None][0]
                        employe.function=worksheet[f"R{row}"].value
                        full_account=int(worksheet[f"S{row}"].value)
                        employe.account_key=int(full_account%100)
                        full_account=int(full_account/100)
                        employe.account_number=int(full_account%100000000)
                        full_account=int(full_account/100000000)
                        employe.account_agency=int(full_account%100000)
                        employe.cni_number=[int(worksheet[f"T{row}"].value) if worksheet[f"T{row}"].value else None][0]
                        employe.cni_established_date=date_handling(worksheet[f"U{row}"].value)
                        employe.cni_established_by=worksheet[f"V{row}"].value
                        if date_handling(worksheet[f"X{row}"].value):
                            partner=Partner()
                            partner.marriage_date=date_handling(worksheet[f"X{row}"].value)
                            partner.name=worksheet[f"Y{row}"].value
                            partner.id_employe=employe
                        employe.unite_id=int(worksheet[f"AD{row}"].value)
                        employe.exit_date=date_handling(worksheet[f"AC{row}"].value)
                        if employe.exit_date is not None:
                            employe.active=3
                        else:
                            print(employe.id,"is active")
                        employe.save()
                    row+=1
        init_code_annual()        
        